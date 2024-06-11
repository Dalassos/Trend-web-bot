import os
import builtins
import re
import json

import datetime

import pandas as pd
from openpyxl import load_workbook

import requests
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.select import Select

from bs4 import BeautifulSoup

import tkinter as tk
from tkinter import filedialog
from tkinter import ttk


#____________________________________________
#Constants

EXCEL_FILE = 'OS_full_list_20240611.xlsx'
SHEET_NAME = 'Trend_OS_full_list'
ACCEPTEDIP = 'ip_list.xlsx'
OUTPUT = 'scan_results_'+str(datetime.datetime.now()).split('.')[0].replace(' ','').replace(':','').replace('-','')+'.xlsx'

#______________________________________________


#functions
def login(text):
    global log
    log.writelines(str(datetime.datetime.now())+" : "+str(text)+"\r")
    print(str(text))

def log_init(out):
    out.write("\r")
    out.write("new log started "+str(datetime.datetime.now()))
    out.write("************************************\r")

def find_origVal(soup, target_text):
    #used for editable fields
    login(f"find_origVal fct, looking for {target_text}")
    try:
        val = find_associated_element(soup, target_text)
        login(f"looking in {val}")
        origVal=val.find('input', attrs={"name":re.compile("origVal$")})['value']
        login(f"Found origVal : {origVal}")
        return origVal
    except Exception as e:
        login(f"Could not find origVal - {e}")

def write_newVal(html_content, target_text, value, driver):
    login(f"write_newVal fct, looking for {target_text}")
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        val = find_associated_element(soup, target_text)
        login(f"looking in {val}")
        newVal_el=val.find('input', attrs={"name":re.compile("newVal$")})
        login(f"Found newVal el: {newVal_el}")
        newVal_id = newVal_el['id']
        login(f"newVal id: {newVal_id}")
        newVal = driver.find_element("id", newVal_id)
        login(f"newVal: {newVal}")
        newVal.clear()
        newVal.send_keys(value)
        return True
    except Exception as e:
        login(f"Could not write newVal - {e}")
        return False
    
def select_newVal(html_content, target_text, value, driver):
    login(f"select_newVal fct, looking for {target_text}")
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        val = find_associated_element(soup, target_text)
        login(f"looking in {val}")
        newVal_el=val.find('select', attrs={"name":re.compile("newVal$")})
        login(f"Found newVal el: {newVal_el}")
        newVal_id = newVal_el['id']
        login(f"newVal id: {newVal_id}")
        newVal = driver.find_element("id", newVal_id)
        login(f"newVal: {newVal}")
        newVal=Select(newVal)
        newVal.select_by_visible_text(value)
        return True
    except Exception as e:
        login(f"Could not select newVal - {e}")
        return False
    
def submit(driver):
    login(f"submit fct")
    try:
        submit_button = driver.find_element("xpath", '//input[@type="image" and @src="images/send.gif"]')
        submit_button.click()
        return True
    except Exception as e:
        login(f"Could not submit - {e}")
        return False

def find_associated_element(soup, target_text):
    #used for non editable fields
    login(f"find_associated_element fct, looking for {target_text}")
    #login(f"in {soup}")
    try:
        # Find all <td> elements with class="pName"
        p_name_elements = soup.find_all('td', class_='pName')
        login(f"p_name_elements: {p_name_elements}")
        
        # Iterate over the <td> elements with class="pName"
        for p_name_element in p_name_elements:
            # Check if the text content matches the target_text
            if p_name_element.get_text(strip=True) == target_text:
                # Get the next sibling <td> element with class="pValue"
                p_value_element = p_name_element.find_next_sibling('td', class_='pValue')
                if p_value_element:
                    login(f"find_associated_element success, returning {p_value_element}")
                    return p_value_element
    except Exception as e:
        login(f"Could not find associated element - {e}")

def visit_webpage_selenium(url, driver):
    login("visit_webpage_selenium fct")
    try:
        # Navigate to the webpage
        driver.get(f"http://{url}")
        page_source = driver.page_source

        # Check if the page loaded successfully
        if ("404 Not Found" not in page_source and page_source != None and page_source !=""):
            login(f"Successfully visited {url}")
            print("Content:")
            print(page_source)
            return page_source, True
        else:
            login(f"Failed to visit {url}. Page not found.")
            return "no answer", False
    except TimeoutException as e:
        login(f"Timeout visiting {url}: {e}")
        return "timeout error", False
    
    except WebDriverException as e:
            login(f"Web driver exception: {url}: {e}")
            return "access error", False
    
def get_alm_dest(ip_address, driver):
    login("get_alm_dest fct")
    try:
        html_content = visit_webpage_selenium(f"{ip_address}/e.htm?ovrideStart=0", driver)  # Replace with the IP address you want to visit
        if (html_content[1] == False) : return "visit error", False
        soup = BeautifulSoup(html_content[0], 'html.parser')
        login("content cast into soup")
        try:
            mainContent = soup.find(id=['mainContent','maindata'])
        except Exception as e:
                login(f"Could not find subdivider: {e}")
        links = mainContent.find_all('a')
        login(str(links))
        return links, True
    except Exception as e:
        login(f"get_alm_dest error: Error getting alarm destinations: {e}")
        return "get_alm_dest error", False

def get_links(url, driver):
    login("get_links fct")
    try:
        html_content = visit_webpage_selenium(f"{url}", driver)
        if (html_content[1] == False) : return "visit error", False
        soup = BeautifulSoup(html_content[0], 'html.parser')
        login("content cast into soup")
        try:
            mainContent = soup.find(id=['mainContent','maindata'])
        except Exception as e:
                login(f"Could not find subdivider: {e}")
        ablocks = mainContent.find_all('a')
        links = []
        for link in ablocks:
            links.append(link.get('href'))
        login(f"links: {links}")
        return links, True
    except Exception as e:
        login(f"get_links error: Error getting alarm destinations: {e}")
        return "get_links error", False

def get_all_pages(ip_address, driver):
    login("get_all_pages fct")
    try:
        html_content = visit_webpage_selenium(f"{ip_address}/modules.htm", driver)  # Replace with the IP address you want to visit
        if (html_content[1] == False) : return "visit error", False
        soup = BeautifulSoup(html_content[0], 'html.parser')
        login("content cast into soup")
        try:
            mainTable = soup.find(class_=('sideMenu'))
            login(f"mainTable: {mainTable}")
        except Exception as e:
                login(f"Could not find subdivider: {e}")
        links = mainTable.find_all('a')
        login(f"links: {links}")
        pages = dict()
        for link in links:
            login(f"link {links.index(link)}: {link}")
            pages.update({link.string : link.get('href')})
        login(f"pages :{pages}")
        return pages, True
    except Exception as e:
        login(f"get_all_pages error: Error getting controller pages: {e}")
        return "get_all_pages error", False
    
def get_time_master_status(ip_address, driver):
    login("get_time Master status fct")
    try:
        html_content = visit_webpage_selenium(f"{ip_address}/T1.htm?ovrideNav:T=FALSE", driver)  # Replace with the IP address you want to visit
        if (html_content[1] == False) : return "visit error", False
        soup = BeautifulSoup(html_content[0], 'html.parser')
        login("content cast into soup")
        timeMasterStatus, manual = scrape_element(soup, "Time Master")
        return timeMasterStatus, manual
    except Exception as e:
        login(f"get_time_Master error: Error getting time Master status: {e}")
        return "get_time_Master error", True
    
def open_alm_dest(html_content):
    login("open_alm_dest fct")
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        login("content cast into soup")
        destination, manual = scrape_dest(soup)
        return destination, manual
    except Exception as e:
        login(f"open_alm_dest error: alarm destination could not be read back: {e}")
        return("open_alm_dest error"), True

def scrape_dest(soup):
    login("scrape_dest fct")
    type = ""
    try:
        type = find_associated_element(soup, "Type").get_text(strip=True)
        login(f"Found type: {type}")
        if (type == "IQ Lan"):
            login("IQ Lan alarm found")
            add = find_origVal(soup, "Address")
            lan = find_origVal(soup, "LAN")
            destination = f"IQLan: L{lan}OS{add}"
            login(f"IQ Lan destination is: {destination}")
            return destination, True
        if (type == "IP"):
            login("IP alarm found")
            dest = find_origVal(soup, "Destination")
            destination = f"{dest}"
            login(f"IP destination is: {destination}")
            return destination, False
    except Exception as e:
        login(f"scrape_dest error: {e}")
        return "error", True

def scrape_element(soup, element):
    login("scrape_element fct")
    value = ""
    error = True
    try:
        value = find_origVal(soup, element)
    except Exception as e:
        login(f"scrape_dest error for origVal: {e} - this can be normal if not editable field")
    if value == None:
        try:
            value = find_associated_element(soup, element).get_text(strip=True)
            login(f"scrape_element success")
        except Exception as e:
            login(f"scrape_dest error for associated_element: {e}")
            value = "error"
            error = False
    return value, error
        
def scrape_page(url, driver, count, max_count):
    login("scrape_page fct")
    try:
        html_content = visit_webpage_selenium(f"{url}", driver) 
        if (html_content[1] == False) : return "visit error", False
        soup = BeautifulSoup(html_content[0], 'html.parser')
        login("content cast into soup")
        #login(f"soup is: {soup}")
        try:
            parameter_table = soup.find(lambda tag: tag.get('id')=='parameterTable' or tag.get('name')=='Adjust')
            login(f"parameter table is: {parameter_table}")
            parameters = parameter_table.find_all('td', class_='pName')
        except Exception as e:
            login(f"Could not find subdividers: {e}")
        login(f"parameter_table: {parameters}")        
        all_fields = dict()
        for parameter in parameters:
            login(f"parameter {parameters.index(parameter)}: {parameter}")
            field = parameter.string
            value, success = scrape_element(parameter_table, parameter.string)
            login(f"parameter {parameters.index(parameter)}: {field} = {value}")
            all_fields.update({field : value})
        return all_fields, success
    except Exception as e:
        login(f"scrape_page error: {e}")
        values = []
        if count < max_count:
            count += 1
            links, success = get_links(url, driver)
            url = url.split("/")[0]
            for sub in links:
                login(f"sublink is: {sub}")
                link = f"{url}/{sub}".replace("//","/")
                value, success = scrape_page(link, driver, count, max_count)
                success *= success
                values.append(value)
            login(f"values : {values}")
            return values, success
        else:
            return "scrape maximum number of sub pages", False

def scrape_all(url,driver,pages_list):
    login(f"scrape_all fct - properties: {pages_list}")
    try:
        with open('json_dat/pages.json', 'r') as json_file:
            pages = json.load(json_file)
            scrape_res=[]
            login(f"pages json loaded: {pages}")
            for page, link in pages.items():
                if page in pages_list:
                    login(f"scraping page: {page}")
                    all_fields = dict()
                    all_fields, success = scrape_page(f"{url}/{link}".replace("//","/"), driver, 0, 2)
                    login(f"scrape_all result: {all_fields}")
                    #for sub in all_fields:
                    #    scrape_res.append([page, sub])
                    #    login(f"scrape_res: {page} : {sub}")
                    scrape_res.append([page, all_fields])
                    login(f"scrape_res: {page} : {all_fields}")
                    #with open(f"json_dat/{page}.json", 'w') as f:
                    #    json.dump(all_fields, f)
            return scrape_res, True
    except Exception as e:
        login(f"scrape_all error: {e}")
        return "error", False

def open_xls(xls, sheetname = None):
    login("opening : "+str(xls))
    try:
        data = pd.read_excel(xls, SHEET_NAME=sheetname)
        return data
    except Exception as e:
        login(f"excel sheet not available: {e}")

def get_column_number(sheet, target_value):
    login(f"get_column_number function, target value: {target_value} in sheet: {sheet}")
    # Iterate over cells in the first row of the sheet
    try:
        for cell in sheet[1]:
            # Check if the cell value matches the target value
            if cell.value == target_value:
                # Return the column number (index) of the matching cell
                result = cell.column
                login(f"get_column_number function complete, column = {result}")
                return result
        login(f"get_column_number function failed, column not found")
    except Exception as e:
        login(f"get_column_number function error : {e}")

def create_xls_prop_sheet(excel_list, property, property_list):
    login(f"create_xls_prop_sheet function")
    try:
        excel_list.create_sheet(property)
        excel_list[property].cell(row=1, column=1).value='site'
        excel_list[property].cell(row=1, column=2).value='Lan'
        excel_list[property].cell(row=1, column=3).value='OS'
        for subprop in property_list:
            i = property_list.index(subprop)
            excel_list[property].cell(row=1, column=i+5).value=clean_prop_name(subprop)
        login(f"create_xls_prop_sheet function complete")
    except Exception as e:
        login(f"create_xls_prop_sheet function error : {e}")

def update_xls_prop_sheet(controller, res, xls, index):
    login(f"update_xls_prop_sheet function")
    try:
        page = res[0]
        props = res[1]
        login(f"page : {page}, props : {props}, index : {index}")
        props.update({'site':controller.site,'Lan':controller.lan,'OS':controller.os})
        for property in props:
            login(f"for property {property}, value is {props[property]}")
            xls[page].cell(row=index, column=get_column_number(xls[page], property)).value = props[property]
        login(f"update_xls_prop_sheet function complete")
    except Exception as e:
        login(f"update_xls_prop_sheet function error : {e}")

def clean_prop_name(prop):
    login(f"clean_prop_name fct init, prop: {prop}")
    try:
        result = re.sub("[{}]","",prop)
        login(f"clean_prop_name fct result: {result}")
        return result
    except Exception as e:
        login(f"clean_prop_name fct error : {e}")

def init_properties_from_json(property):
    login(f"init_properties_from_json fct init, property: {property}")
    try:
        with open(f"json_dat/{property}.json", 'r') as json_file:
            json_prop = json.load(json_file)
            login(f"json_prop is: {json_prop}")
            property_list = []
            for page in json_prop:
                for item in page.items():
                    prop, value = item
                    login(f"prop is: {prop}, value is: {value}")
                    property_list.append(prop)
            sorted(set(property_list))
            login(f"init_properties_from_json fct success, property_list : {property_list}")
            return property_list
    except Exception as e:
        login(f"init_properties_from_json fct error : {e}")

class controller:
    def __init__(self, row):
        self.site = row['siteLabel']
        self.lan = row['LanNo']
        self.os = row['NodeAddress']
        self.ip = row['nodeIpAddr']

#____________________________________________
#GUI functions

class GUI:
    def cancel(self):
        quit()

    def scan(self):
        #global Replace 
        Confirm = False

        def execute(self):
            if Replace == True:
                origin_ip = t963Ip
                final_ip = '111.111.111.111'
                final_format = format_IQVision

            #testing shortcut
            #self.selected_sites = ['Lister LTC']
            #self.selected_properties = ['Address Page', 'Performance']

            login(f"sites to action: {self.selected_sites}")
            login(f"pages to read: {self.selected_properties}")

            #list properties
            excel_list = load_workbook(EXCEL_FILE)
            for property in self.selected_properties:
                property_list = init_properties_from_json(property)
                create_xls_prop_sheet(excel_list, property, property_list)

            # Initialize the WebDriver (replace 'chromedriver' with the path to your driver executable)
            options = webdriver.ChromeOptions()
            options.add_argument('ignore-certificate-errors')
            options.add_argument('acceptInsecureCerts')
            with webdriver.Chrome(options=options) as driver:
                try:
                    # Loop through all rows using iterrows()
                    scrape_index = 1
                    for index, row in os_list.iterrows():
                        try:
                            manual = False
                            visit_success = True
                            login("new row of excel sheet")
                            TrendCont = controller(row)
                            #TrendCont.ip = '172.16.7.195'
                            if TrendCont.site in self.selected_sites and (TrendCont.os != 126):
                                login(f"controller to check: {TrendCont.site} - {TrendCont.ip}")
                                if (TrendCont.ip == "#N/A#" or TrendCont.ip == "" or TrendCont.ip == "128.1.1.3" or TrendCont.ip == "inv" or pd.isna(TrendCont.ip)):
                                    visit_success = False
                                    login(f"controller not visitable : {TrendCont.site} - {TrendCont.ip}")
                                else:
                                    alm_dest, visit_success = get_alm_dest(TrendCont.ip, driver)
                                if (visit_success == False) :
                                        login("no access to this controller")
                                else :
                                    scrape_index += 1
                                    scrape_res, success = scrape_all(TrendCont.ip, driver, self.selected_properties)
                                    login(f"scrape_res for scrape all: {scrape_res}")
                                    for res in scrape_res:
                                        login(f"res is: {res}")
                                        update_xls_prop_sheet(TrendCont, res, excel_list, scrape_index)
                                    try :
                                        alm_dest.length()
                                    except :
                                        manual = True
                                    else :
                                        #only if we require multiple connections together
                                        manual = (alm_dest.lenght()<1)
                                        #manual = False 
                                    for alm in alm_dest :
                                        i = alm_dest.index(alm)
                                        #past_alm_dest = str(row["Alarm Destinations e"+str(i+1)])
                                        href_value = alm.get('href')
                                        #login(f"known alarm dest {href_value} is: {past_alm_dest}")
                                        url = f"{TrendCont.ip}/{href_value}".replace("//","/")
                                        login(f"alarm at this url: {url}")
                                        html_content, visit_success = visit_webpage_selenium(url, driver)
                                        destination, instance_manual = open_alm_dest(html_content)
                                        network, instance_manual = visit_webpage_selenium(f'{url}/n.htm?ovrideStart=0', driver)
                                        manual *= instance_manual
                                        if visit_success == True:
                                            out.write("e"+str(i)+" destination : "+str(destination)+" \r")
                                            column = get_column_number(excel_list[SHEET_NAME], f'Alarm Destinations e{i + 1}')
                                            login(f"writing to column {column}")
                                            excel_list[SHEET_NAME].cell(row=index+2, column=column).value = destination
                                            if Replace == True:
                                                if destination == origin_ip :
                                                    #insert code to overwrite the former IP address
                                                    select_newVal(html_content, "Message Format", final_format, driver)
                                                    write_newVal(html_content, "Destination", final_ip, driver)
                                                    submit(driver)
                                if (manual == True or visit_success == False) :
                                    login(f"Controller will require manual intervention")
                                    error.writelines(f"{TrendCont.site} - {TrendCont.ip} will require manual intervention")
                                    column = 15
                                    excel_list[SHEET_NAME].cell(row=index+2, column=column).value="Manual Intervention Required"
                                    if Replace == True:
                                        error.writelines(f"{TrendCont.site} - {TrendCont.ip} could not be updated")
                        except Exception as e:
                            login(f"Controller failure, skipping controller {row} - "+str(e))
                            error.writelines(f"{TrendCont.site} - {TrendCont.ip} could not be accessed")
                    try:
                        #add in code for custom output in case of failure
                        excel_list.save(OUTPUT)
                    except Exception as e:
                        login(f"Write error: {e}")
                except Exception as e:
                    login("Major failure, exiting now - "+str(e))
                driver.close()
                login("Done")

        def confirm_replace():
            nonlocal popup
            Confirm = True
            popup.destroy
            execute(self)

        if Confirm == False and Replace == True :
            popup = tk.Toplevel(self.root)
            popup.title("Replace")

            confirm_button = tk.Button(popup, text="Confirm", command=confirm_replace)
            confirm_button.pack()
            cancel_button = tk.Button(popup, text="Cancel", command=self.cancel)
            cancel_button.pack()
            confirm_text = tk.Label(popup, text = "Are you sure you want to write to controllers?",wraplength=150, width=35, height=15)
            confirm_text.pack()

        else:
            execute(self)

    def createChkbx(self):
        return GUI.checkbox_list(self)
    
    class checkbox_list:

        def __init__(self, outer_instance):
            self.outer_instance = outer_instance
            self.checkboxes = []
            self.selected_values = []

        def show_checkbox_list(self, ckb_list):
            
            def confirm_selection(self):
                self.selected_values = [item for item, var in self.checkboxes if var.get()]
                popup.destroy()

            def toggle_select_all(self):
                select_all_state = select_all_var.get()
                for var in self.checkboxes:
                    #login(f"var is: {var[1]}")
                    var[1].set(select_all_state)
            
            popup = tk.Toplevel(self.outer_instance.root)
            popup.title("Checkbox List")
            

            # Create a variable for "Select All" checkbox
            select_all_var = tk.BooleanVar()
            select_all_var.set(False)  # Initially not selected

            # Calculate number of columns based on the number of options
            num_columns = 3
            num_options = len(ckb_list)+1
            num_rows = -(-num_options // num_columns)  # Equivalent to math.ceil(num_options / num_columns)

            # Create the "Select All" checkbox
            select_all_checkbox = tk.Checkbutton(popup, text="Select All", variable=select_all_var, command=lambda: toggle_select_all(self))
            select_all_checkbox.grid(row=0, column=0, sticky="w")

            for i, item in enumerate(ckb_list):
                row = (i+1) // num_columns
                column = (i+1) % num_columns
                var = tk.BooleanVar()
                checkbtn = tk.Checkbutton(popup, text=item, variable=var)
                checkbtn.grid(row=row, column=column, sticky="w")
                self.checkboxes.append((item, var))

            confirm_button = tk.Button(popup, text="Confirm", command=lambda: confirm_selection(self))
            confirm_button.grid(row=num_rows, columnspan=num_columns, pady=10)

            popup.grab_set()  # Make the popup modal
            popup.wait_window()  # Wait for the popup window to close
            login(f"selected values : {self.selected_values}")
            
            return self.selected_values

    def select_property(self, property_list):
       property_chkbx = self.checkbox_list(self)
       login(f"self.selected_properties: {self.selected_properties}")
       self.selected_properties = property_chkbx.show_checkbox_list(property_list)
       login(f"self.selected_properties: {self.selected_properties}")

    def select_sites(self, sites_list):
        sites_chkbx = self.checkbox_list(self)
        login(f"self.selected_sites: {self.selected_sites}")
        self.selected_sites = sites_chkbx.show_checkbox_list(sites_list)
        login(f"self.selected_sites: {self.selected_sites}")

    def on_checkbox_toggle(self):
            global Replace 
            Replace = not Replace
            login(f"Replace mode = {Replace}")
            return Replace

    def __init__(self, sites_list, property_list):
        global Replace
        self.sites_list = sites_list
        self.property_list = property_list
        self.selected_sites = []
        self.selected_properties = []

        #Tkinter GUI
        root = tk.Tk()
        root.title("Trend Alarm destination crawler")
        root.minsize(800,200)
        root.geometry("480x100")

        
        # create the main sections of the layout, 
        # and lay them out
        buffer = tk.Frame(root, width=200, height=20)
        top = tk.Frame(root)
        middle = tk.Frame(root)
        bottom = tk.Frame(root)
        buffer.pack(side=tk.TOP)
        top.pack(side=tk.TOP)
        middle.pack(side=tk.TOP)
        bottom.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        buffer.pack(side=tk.TOP)


        # create the widgets for the top part of the GUI,
        # and lay them out
        s = tk.Button(root, text="Select sites", width=10, height=2, command=lambda: self.select_sites(self.sites_list))
        p = tk.Button(root, text="Select property to scan", width=10, height=2, command=lambda: self.select_property(self.property_list))
        c = tk.Button(root, text="Leave", width=10, height=2, command=self.cancel)
        e = tk.Button(root, text="Scan", wraplength=60, width=10, height=2, command=self.scan)
        #r = tk.Checkbutton(root, text="Replace", variable=Replace, width=10, height=2, command=self.on_checkbox_toggle)
        #r.pack(in_=top, side=tk.LEFT)
        s.pack(in_=middle, side=tk.LEFT)
        p.pack(in_=middle, side=tk.LEFT)
        e.pack(in_=middle, side=tk.LEFT)
        c.pack(in_=middle, side=tk.LEFT)

        # create the widgets for the bottom part of the GUI,
        # and lay them out
        global path
        path = tk.Label(root, text = "Select 'Scan' to read selected files and check the replace box to make replacements", width=35, height=15)
        path.pack(in_=bottom, side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        w = tk.Label(root, text="Please choose SET Project file directory")
        w.pack()

        self.root = root
        

#__________________________________________________________________________
# main 
with open("trend_web_bot.log","w") as log, open("error.log","w") as error, open("alarmDest.log","w") as out:
    log_init(log)
    log_init(error)

    # Load the Excel file
    try:
        os_list = pd.read_excel(EXCEL_FILE, SHEET_NAME)
        ip_list = pd.read_excel(ACCEPTEDIP)
    except Exception as e:
        login(f"error reading spreadsheets: {e}")

    #make sites list
    try:
        unique_Sites = sorted(set(os_list['siteLabel'].dropna()))
        login(f"sites list : {unique_Sites}")
    except Exception as e:
        login(f"error creating site list: {e}")

    #make property list
    try:
        with open('json_dat/pages.json', 'r') as json_file:
            pages = json.load(json_file)
            property_list = []
            for page, link in pages.items():
                property_list.append(page)
            sorted(set(property_list))
            login(f"property_list : {property_list}")
    except Exception as e:
        login(f"error creating property_list: {e}")

        

    #read listed IPs
    try:
        for index, row in ip_list.iterrows():
            if index == 0:
                IQVisionIp = row['IQVision']
                t963Ip = row['t963']
            if index == 1:
                format_963 = row['t963']
                format_IQVision = row['IQVision']
        login(f"t963: {t963Ip} - {format_963}, Vision: {IQVisionIp} - {format_IQVision}")
    
    except Exception as e:
        login(f"error getting IPs: {e}")

    #init mode
    Replace = False

    checkbox_var = False
    gui = GUI(unique_Sites, property_list)
    gui.root.mainloop()
    